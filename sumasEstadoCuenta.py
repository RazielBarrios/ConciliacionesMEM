import pandas as pd
from openpyxl import load_workbook

def sumasEC():
    archivo = "conciliacion.xlsx"
    filesheet = "conciliacion.xlsx"

    wb = load_workbook(filesheet)
    hojaEscritura = wb.worksheets[1]
    
    #Lee el archivo de datos de origen
    hojaLectura = pd.read_excel(io = archivo, sheet_name="Resumen", header = None)
    
    #Suma de las columnas P a V
    suma = 0
    aux = chequeo(hojaEscritura, 16) - 2
    for i in (16, 18, 19, 20, 21):
        suma += float(hojaLectura[i][aux])
    print(suma)
    fil = chequeo(hojaEscritura, 24)
    hojaEscritura.cell(row=fil, column=24).value = str(suma)
    
    #Suma de las columas Z y AB
    suma = 0
    aux = chequeo(hojaEscritura, 25) - 2
    print(hojaLectura[25][aux])
    print(hojaLectura[27][aux])
    suma = float(hojaLectura[25][aux]) + float(hojaLectura[27][aux])
    print(suma)
    fil = chequeo(hojaEscritura, 29)
    hojaEscritura.cell(row=fil, column=29).value = str(suma)
    
    wb.save('conciliacion.xlsx')
    
def chequeo(sheet, b):
    a=1
    celda = sheet.cell(row=a, column=b)
    while celda.value:
        a=a+1
        celda = sheet.cell(row=a, column=b)
    return a

#sumasEC()