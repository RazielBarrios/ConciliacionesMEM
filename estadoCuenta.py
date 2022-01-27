from xml.etree import ElementTree
from xml.dom import minidom
from openpyxl import load_workbook
from notasDebitoCredito import *
from sumasEstadoCuenta import *

def extEstadoCuenta():
    #Apertura del archivo para escritura de informacion
    filesheet = "conciliacion.xlsx"
    wb1 = load_workbook(filesheet)
    hojaEscritura = wb1.worksheets[1]
    estCuenta = 'estadodeCuenta.xml'
    #Apertura de archivos para extraccion de informacion
    tree = ElementTree.parse('estadodeCuenta.xml')
    root = tree.getroot()
    suma = 0
    doc = minidom.parse('estadodeCuenta.xml')
    
    #Se obtiene el identificador del estado de cuenta, pues con cada nuvo estado de cuensta este cambia
    ident = doc.getElementsByTagName('estadodecuenta')
    identificador = ident[0].attributes['FUECD'].value
    
    importe = doc.getElementsByTagName('monto_total')
    
    #se obtiene la suma de los MWh para la columna P del resumen de estado de cuenta
    ID = "'" + identificador + "-B02020'" 
    t = 'POTENCIA_MDA' 
    s = calculoMWH(ID, root, t)
    escrituraDatos(hojaEscritura, 16, s)
    
    #Se obtiene el importe para la columna Q del resumen del estado de cuenta
    ID = identificador + "-B02020" 
    dato = obtenerMonto(doc, ID)
    escrituraDatos(hojaEscritura, 17, dato)
      
    #Llenado de la columna R  
    ID = "'" + identificador + "-C50120'" 
    t = 'POTENCIA' 
    s = calculoMWH(ID, root, t)   
    escrituraDatos(hojaEscritura, 18, s)
    
    #Llenado de la columna S
    ID = identificador + "-C54120" 
    dato = obtenerMonto(doc, ID)
    escrituraDatos(hojaEscritura, 19, dato)
    
    #Llenado de la columna T        
    ID = identificador + "-C50120" 
    dato = obtenerMonto(doc, ID)
    escrituraDatos(hojaEscritura, 20, dato)
    
    #Llenado columna X
    
    #Lenado de la columna Y
    ID = "'" + identificador + "-A01010'" 
    t = 'POTENCIA_MDA' 
    s = calculoMWH(ID, root, t)    
    escrituraDatos(hojaEscritura, 25, s)
    
    #Llenado de la columna Z
    ID = identificador + "-A01010" 
    dato = obtenerMonto(doc, ID)
    escrituraDatos(hojaEscritura, 26, dato)
    
    #Lenado de la columna AA
    ID = "'" + identificador + "-B01010'" 
    t = 'POTENCIA_MTR' 
    s = calculoMWH(ID, root, t)    
    escrituraDatos(hojaEscritura, 27, s)
    
    #Llenado de la columna AB
    ID = identificador + "-B01010" 
    dato = obtenerMonto(doc, ID)
    escrituraDatos(hojaEscritura, 28, dato)
    
    wb1.save('conciliacion.xlsx')
    
    extDebitoCredito(filesheet, estCuenta)
    sumasEC()


def calculoMWH(iD, rt, tipo):
    suma = 0
    for node in rt.findall("./liquidaciones/liquidacion/facturas/factura/conceptos/concepto/[@ful=" + iD + "]/anexos/anexo/registroshorarios/registro"):
        #print(node.attrib.get('HORA'), node.attrib.get(tipo))
        suma += float(node.attrib.get(tipo))
    return suma
    
def obtenerMonto(docto, ident):
    importe = docto.getElementsByTagName('monto_total')  
    for i in range (len(importe)):     #para condicionar la busqueda a un
        concepto = docto.getElementsByTagName('concepto')
        condicion = concepto[i].attributes['ful'].value #valor en específico
        if condicion == ident:     #Condicion de busqueda
            dt = importe[i].firstChild.data      #Se extrae el dato
            print(dt)
            return dt
            
def escrituraDatos(hEscritura, colEscritura, dato):
    a=1
    celda = hEscritura.cell(row=a, column=colEscritura)
    while celda.value:
        a=a+1
        celda = hEscritura.cell(row=a, column=colEscritura)
    hEscritura.cell(row=a, column=colEscritura).value = str(dato)
    
#extEstadoCuenta()