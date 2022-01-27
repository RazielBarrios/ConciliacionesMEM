from openpyxl import load_workbook
import pandas as pd


def calculoCM(archivo,b):
    
    #archivo = input("Seleccione archivo cincominutal MDA no generado:")
    #Lee el archivo de datos cincominutales
    df = pd.read_excel(io = archivo, sheet_name=0, skiprows = 1, header = None, usecols = [2])
    filesheet = "conciliacion.xlsx"
    tam = len(df)

    #Apertura del archivo destino
    wb = load_workbook(filesheet)
    sheet = wb.worksheets[0]
 
    #Escritura de los datos del archivo 1 al archivo 2
    colCincominutal = 2
    valorHora = 0
    contadorHoras = 1
    a=1
    #b=3
    celda = sheet.cell(row=a, column=b)
    while celda.value:
        a=a+1
        celda = sheet.cell(row=a, column=b)
    rowHora = a
    rowCincominutal = 0
    while rowCincominutal < tam:
        if contadorHoras <= 12:
            valorHora = valorHora + float(df[colCincominutal][rowCincominutal])
            contadorHoras = contadorHoras + 1
            rowCincominutal = rowCincominutal + 1
        else:
            sheet.cell(row=rowHora, column=b).value = str(valorHora/1000)
            print(valorHora)
            rowHora = rowHora + 1
            contadorHoras = 1
            valorHora = 0
    sheet.cell(row=rowHora, column=b).value = str(valorHora/1000)
    wb.save('conciliacion.xlsx')
 