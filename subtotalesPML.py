import pandas as pd
from openpyxl import load_workbook

def multColumnas(colMDA, colMTR, PMDA, PMTR, b1, b2):
    archivo = "conciliacion.xlsx"

    filesheet = "conciliacion.xlsx"
     
    wb = load_workbook(filesheet)
    sheet = wb.worksheets[0]
    #Lee el archivo de datos de origen
    df = pd.read_excel(io = archivo, sheet_name="Precio", header = None)

   
    for n in range(1,5):
        a=1
        celda = sheet.cell(row=a, column=b1)
        while celda.value:
            a=a+1
            celda = sheet.cell(row=a, column=b1)
        i = a
        for i in range(1,25):
            sheet.cell(row=a, column=b1).value = str(float(df[colMDA][i]) * float(df[PMDA][i]))
            a = a + 1
        
        b1+=1
        PMDA+=1
    n=1    
    for n in range(1,5):
        a=1
        celda = sheet.cell(row=a, column=b2)
        while celda.value:
            a=a+1
            celda = sheet.cell(row=a, column=b2)
        i = a
        for i in range(1,25):
            sheet.cell(row=a, column=b2).value = str(float(df[colMTR][i]) * float(df[PMTR][i]))
            a = a + 1
        
        b2+=1
        PMTR+=1
        
    wb.save('conciliacion.xlsx')

#multColumnas(3, 4, 7, 11, 16, 20)