import pandas as pd
from openpyxl import load_workbook
from estadoCuenta import extEstadoCuenta

def generacionResumen():
    archivo = "conciliacion.xlsx"
    eCuenta = 'estadodeCuenta.xml'
    wb = load_workbook(archivo)
    hojaEscritura = wb.worksheets[1]
    #Lee el archivo de datos de origen
    hojaLectura = pd.read_excel(io = archivo, sheet_name="Precio", header = None)
    hojaLectura2 = pd.read_excel(io = archivo, sheet_name="Tarifas", header = None)
    #hojaLectura3 = pd.read_excel(io = archivo, sheet_name="Resumen", header = None)

    #Llenado de las Fechas
    
    
    a2 = chequeo(hojaEscritura, 1)
    fecha = int(hojaLectura2[1][2]) - 1
    hojaEscritura.cell(row=a2, column=1).value = str(hojaLectura[0][fecha])
    
      
    #Llenado de la columna B, sumatoria de las 24 hrs de "Medicion de energía MDA" (columna C, hoja "precios")
    hacerSumaB(2,2,hojaEscritura, hojaLectura, hojaLectura2)
    #Llenado de la columna C, sumatoria de las 24 hrs de "Energía MDA" (columna D, hoja "precios")
    hacerSuma(3,3,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(3,3, hojaEscritura)
    #Llenado de la columna D, sumatoria de las 24 hrs de "Energía MDA" (columna P, hoja "precios")
    hacerSuma(15,4,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(4,4, hojaEscritura)
    #Llenado de la columna E, sumatoria de las 24 hrs de "Energía MDA" (columna Q, hoja "precios")
    hacerSuma(16,5,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(5,5, hojaEscritura)
    #Llenado de la columna F, sumatoria de las 24 hrs de "Energía MDA" (columna R, hoja "precios")
    hacerSuma(17,6,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(6,6, hojaEscritura)
    #Llenado de la columna G, sumatoria de las 24 hrs de "Energía MDA" (columna S, hoja "precios")
    hacerSuma(18,7,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(7,7, hojaEscritura)
    
    #Llenado de la columna H, sumatoria de las 24 hrs de "Energía MTR" (columna E, hoja "precios")
    hacerSuma(4,8,hojaEscritura, hojaLectura, hojaLectura2)
    #Llenado de la columna I, sumatoria de las 24 hrs de "Energía MTR" (columna T, hoja "precios")
    hacerSuma(19,9,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(9,9, hojaEscritura)
    #Llenado de la columna J, sumatoria de las 24 hrs de "Energía MTR" (columna U, hoja "precios")
    hacerSuma(20,10,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(10,10, hojaEscritura)
    #Llenado de la columna K, sumatoria de las 24 hrs de "Energía MTR" (columna V, hoja "precios")
    hacerSuma(21,11,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(11,11, hojaEscritura)
    #Llenado de la columna L, sumatoria de las 24 hrs de "Energía MTR" (columna W, hoja "precios")
    hacerSuma(22,12,hojaEscritura, hojaLectura, hojaLectura2)
    signoPesos(12,12, hojaEscritura)
    
    llenarColumnaO(hojaEscritura, 15)
    
     
    wb.save(archivo)   
    
    
    extEstadoCuenta(archivo, eCuenta)
    
    wb = load_workbook(archivo)
    hojaEscritura = wb.worksheets[1]
    for i in (13,14,17,19,20,21,22,24,26,28,29,30,31):
        signoPesos(i,i, hojaEscritura)
    wb.save(archivo)

def chequeo(sheet, b):
    a=1
    celda = sheet.cell(row=a, column=b)
    while celda.value:
        a=a+1
        celda = sheet.cell(row=a, column=b)
    return a

def llenarColumnaO(hEscr, colEscr):
    a1 = chequeo(hEscr, colEscr)
    hEscr.cell(row=a1, column=colEscr).value = "$ 0.00"

def hacerSuma(colLectura, colEscritura, hEscritura, hLectura, hTarifas):
    a1 = chequeo (hEscritura, colEscritura)
    vInicio = int(hTarifas[1][2]) - 24
    vFinal = int(hTarifas[1][2]) - 1
    suma = 0
    for i in range(vInicio, vFinal):
        suma += float(hLectura[colLectura][i])
    hEscritura.cell(row=a1, column=colEscritura).value = str(suma)

def hacerSumaB(colLectura, colEscritura, hoja, df, hoja2):
    a1 = chequeo (hoja, colEscritura)
    suma = 0
    vInicio = int(hoja2[1][2]) - 24
    vFinal = int(hoja2[1][2]) - 1
    for j in range(vInicio, vFinal):
        suma += float(df[colLectura][j])
    hoja.cell(row=a1, column=colEscritura).value = str(suma)
    valor = hoja2[1][0]
    costoMWH = valor.split()
    hoja.cell(row=a1, column=13).value = str(suma * float(costoMWH[1]))
    valor = hoja2[1][1]
    cobroMWH = valor.split()
    hoja.cell(row=a1, column=14).value = str(suma * float(cobroMWH[1]))

           
def signoPesos(colLectura, colEscritura,  hojaE):
    a1 = chequeo (hojaE, colEscritura) - 1
    hojaE.cell(row=a1, column=colEscritura).value = "$ " + str(hojaE.cell(row=a1, column=colEscritura).value)
        

generacionResumen()