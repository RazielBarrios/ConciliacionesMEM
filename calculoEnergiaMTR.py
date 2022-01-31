import pandas as pd
from openpyxl import load_workbook

def restaColumnas(col1, col2, b, archivoCon):
    #archivo = "conciliacion.xlsx"

    #filesheet = "conciliacion.xlsx"
     
    wb = load_workbook(archivoCon)
    sheet = wb.worksheets[0]

    a=1
    celda = sheet.cell(row=a, column=b)
    while celda.value:
        a=a+1
        celda = sheet.cell(row=a, column=b)
    
    #Lee el archivo de datos de origen
    df = pd.read_excel(io = archivoCon, sheet_name="Precio", header = None)
    i = a
    for i in range(1,25):
        sheet.cell(row=a, column=b).value = str(float(df[col1][i]) - float(df[col2][i]))
        a = a + 1
    wb.save(archivoCon)
