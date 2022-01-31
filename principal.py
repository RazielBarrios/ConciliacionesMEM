import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from creacionArchivoConciliacion import creacionArchivoFinal
from calculoCincominutalMDA import calculoCM
from extraccionPuntoBase import puntoBase
from calculoEnergiaMTR import restaColumnas
from extraccionWeb import extraccionW
from subtotalesPML import multColumnas


# Hoja de trabajo de operación, si no existe, se crea
archivoConciliacion = "conciliacion.xlsx"
fileObj = Path(archivoConciliacion)
if fileObj.is_file() == False:
    creacionArchivoFinal()
    
fecha = input("Ingresa la fecha de la conciliacion: ")
    
wb = load_workbook(archivoConciliacion)
sheet = wb.worksheets[0]

#Llenado de fecha y hora, columnas A y B
a=1
b=1
celda = sheet.cell(row=a, column=b)
while celda.value:
    a=a+1
    celda = sheet.cell(row=a, column=b)
        
for i in range (1,25):
    sheet.cell(row=a, column=b).value = fecha
    sheet.cell(row=a, column=b+1).value=i
    a+=1
    b=1
wb.save(archivoConciliacion)

#Lleenado de columna C, valores cincominutales ya calculados a Megawatts/Hora
arch = input("Seleccione archivo cincominutal MDA no generado:")
calculoCM(arch, 3, archivoConciliacion)

#Llenado de columna D, punto base
arch = input("Seleccione archivo de energía asignada:")
puntoBase(arch, archivoConciliacion)

#Llenado de columna E
restaColumnas(2,3,5,archivoConciliacion) #Por el uso de diferentes librerias, para la lectura las columnas se cuentan desde el numero 0 y para la escritura desde el numero 1

#Llenado de columna F
arch = input("Seleccione archivo cincominutal MTR generado:")
calculoCM(arch, 6, archivoConciliacion)

#Llenado de columna G
restaColumnas(5,2,7, archivoConciliacion)


extraccionW(archivoConciliacion, fecha)

multColumnas(3, 4, 7, 11, 16, 20, archivoConciliacion)


#archivo = "conciliacion.xlsx"
hoja= pd.read_excel(io = archivoConciliacion, sheet_name="Precio", header = None, usecols = [0])
a = len(hoja)
#filesheet = "conciliacion.xlsx"
wb = load_workbook(archivoConciliacion)
hojaEscritura = wb.worksheets[2]
hojaEscritura.cell(row=3, column=2).value = str(a)
wb.save(archivoConciliacion)















