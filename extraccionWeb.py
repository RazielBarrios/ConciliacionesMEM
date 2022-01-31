from urllib import request
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime

def extraccionW(archivoCon, fecha):
    fecha= datetime.strptime(fecha, '%d/%m/%Y')
    #url1 = "https://ws01.cenace.gob.mx:8082/SWPML/SIM/SIN/MDA/01VAJ-230/2021/11/26/2021/11/26/XML"
    #url2 = "https://ws01.cenace.gob.mx:8082/SWPML/SIM/SIN/MTR/01VAJ-230/2021/11/26/2021/11/26/XML"
    url1 = "https://ws01.cenace.gob.mx:8082/SWPML/SIM/SIN/MDA/01VAJ-230/" + str(fecha.year) + "/" + str(fecha.month) + "/" + str(fecha.day) + "/" + str(fecha.year) + "/" + str(fecha.month) + "/" + str(fecha.day) + "/XML"
    url2 = "https://ws01.cenace.gob.mx:8082/SWPML/SIM/SIN/MTR/01VAJ-230/" + str(fecha.year) + "/" + str(fecha.month) + "/" + str(fecha.day) + "/" + str(fecha.year) + "/" + str(fecha.month) + "/" + str(fecha.day) + "/XML"
    
    print(url1)
    print(url2)
    html = request.urlopen(url1)
    data = html.read()
    tree = ET.fromstring(data)
    results = tree.findall('Resultados/Nodo/Valores/Valor')
    
   #filesheet = "conciliacion.xlsx"
    wb = load_workbook(archivoCon)
    sheet = wb.worksheets[0]
    a=1
    b=8
    palabra = ""
    for b in range (8,12):
        celda = sheet.cell(row=a, column=b)
        while celda.value:
            a=a+1
            celda = sheet.cell(row=a, column=b)
        if b == 8:
            palabra = 'pml'
        if b == 9:
            palabra = 'pml_ene'
        if b == 10:
            palabra = 'pml_per'
        if b == 11:
            palabra = 'pml_cng'
        for result in results:
            sheet.cell(row=a, column=b).value = result.find(palabra).text
            a += 1
        a = 1
    wb.save(archivoCon)
    
    html = request.urlopen(url2)
    data = html.read()
    tree = ET.fromstring(data)
    results = tree.findall('Resultados/Nodo/Valores/Valor')
    #filesheet = "conciliacion.xlsx"
    wb = load_workbook(archivoCon)
    sheet = wb.worksheets[0]
    a=1
    b=12
    palabra = ""
    for b in range (12,16):
        celda = sheet.cell(row=a, column=b)
        while celda.value:
            a=a+1
            celda = sheet.cell(row=a, column=b)
        if b == 12:
            palabra = 'pml'
        if b == 13:
            palabra = 'pml_ene'
        if b == 14:
            palabra = 'pml_per'
        if b == 15:
            palabra = 'pml_cng'
        for result in results:
            sheet.cell(row=a, column=b).value = result.find(palabra).text
            a += 1
        a = 1
    
    wb.save(archivoCon)

"""archivo = "conciliacion.xlsx"
fech = input("Ingrese una fecha: ")        
extraccionW(archivo, fech)"""