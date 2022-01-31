from xml.etree import ElementTree
from xml.dom import minidom
from openpyxl import load_workbook
from datetime import datetime

def extDebitoCredito(archivoEscr, Axml):
    #Apertura del archivo para escritura de informacion
    #filesheet = "conciliacion.xlsx"
    wb1 = load_workbook(archivoEscr)
    hojaEscritura1 = wb1.worksheets[1]
    hojaEscritura2 = wb1.worksheets[3]
    
    #Apertura de archivos para extraccion de informacion
    tree = ElementTree.parse(Axml)
    root = tree.getroot()
    doc = minidom.parse(Axml)
    
    persona = ['"cenace"', '"participante"']
    tFactura = ['"cobronotadebito"', '"cobronotacredito"', '"pagonotadebito"', '"pagonotacredito"']
    tFinal=["Cenace Débito", "Cenace Crédito", "Participante Débito", "Participante Crédito"]
    s = [0, 0, 0, 0]
    notas = [[0, 0, 0, 0],
             [0, 0, 0, 0],
             [0, 0, 0, 0],
             [0, 0, 0, 0]]
    
    for a in range (1,5):
        b = 0
        liq = str (a)
        liq = '"' + liq + '"'
        for node in root.findall("./liquidaciones/liquidacion/[@num_liq=" + liq + "]/facturas/factura/[@emisor=" + persona[0] + "]/[@tipo= " + tFactura[0] + "]/conceptos/concepto/monto_total_dif"):
            notas [a-1][b] = round(notas[a-1][b] + float(node.text), 2)
            s[0]= round(s[0] + float(node.text), 2)
        b += 1
        for node in root.findall("./liquidaciones/liquidacion/[@num_liq=" + liq + "]/facturas/factura/[@emisor=" + persona[0] + "]/[@tipo= " + tFactura[1] + "]/conceptos/concepto/monto_total_dif"):
            notas [a-1][b] = round(notas[a-1][b] + float(node.text), 2)
            s[1]= round(s[1] + float(node.text), 2)
        b += 1
        for node in root.findall("./liquidaciones/liquidacion/[@num_liq=" + liq + "]/facturas/factura/[@emisor=" + persona[1] + "]/[@tipo= " + tFactura[2] + "]/conceptos/concepto/monto_total_dif"):
            notas [a-1][b] = round(notas[a-1][b] + float(node.text), 2)
            s[2]= round(s[2] + float(node.text), 2)
        b += 1
        for node in root.findall("./liquidaciones/liquidacion/[@num_liq=" + liq + "]/facturas/factura/[@emisor=" + persona[1] + "]/[@tipo= " + tFactura[3] + "]/conceptos/concepto/monto_total_dif"):
            notas [a-1][b] = round(notas[a-1][b] + float(node.text), 2)
            s[3]= round(s[3] + float(node.text), 2)
        b += 1
    
    #Guarda información en la hoja de resumen de estado de cuenta
    for i in range(0,4):
        if s[i]<0:
            s[i] = s[i] * -1
        print(tFinal[i] + ": " + str(s[i]))
        if i == 0:
            escrituraDatos(hojaEscritura1, 30, str(s[i]))
        elif i == 1:
            escrituraDatos(hojaEscritura1, 31, str(s[i]))
        elif i == 2:
            escrituraDatos(hojaEscritura1, 21, str(s[i]))
        elif i == 3:
            escrituraDatos(hojaEscritura1, 22, str(s[i]))
    
    #Guarda información en la hoja de notas de debito y credito
    ident = doc.getElementsByTagName('estadodecuenta')
    fecha = ident[0].attributes['fecha_oper'].value
    fecha= datetime.strptime(fecha, '%Y/%m/%d')
    f = str(fecha.day) + "/" + str(fecha.month) + "/" + str(fecha.year)
    print(f)
    escrituraDatos2(hojaEscritura2, 1, f)
        
    b = 2
    for i in range (0, 4):
        print("Liquidacion " + str(i+1) + ": ")
        for j in range (0, 4):
            print(tFinal[j] + ": " + str(notas[i][j]))
            escrituraDatos2(hojaEscritura2, b, str(notas[j][i]))
            b += 1
        
    wb1.save(archivoEscr)
    
def escrituraDatos(hEscritura, colEscritura, dato):
    a=1
    celda = hEscritura.cell(row=a, column=colEscritura)
    while celda.value:
        a=a+1
        celda = hEscritura.cell(row=a, column=colEscritura)
    hEscritura.cell(row=a, column=colEscritura).value = dato
    
#Se define una funcion extra de scritura de datos por la diferencia de parametros para la verificacion de celdas vacias (variable a), por la
#estructura de la hoja de nostas de debito y credito (el formato de hoja predefinido) es importante aclarar en que fila comienza la escritura del archivo
    
def escrituraDatos2(hEscritura, colEscritura, dato):
    a=4
    celda = hEscritura.cell(row=a, column=colEscritura)
    while celda.value:
        a=a+1
        celda = hEscritura.cell(row=a, column=colEscritura)
    hEscritura.cell(row=a, column=colEscritura).value = dato
    

"""hoja = "conciliacion.xlsx"
archivoXML = 'xml2.xml'
extDebitoCredito(hoja, archivoXML) """