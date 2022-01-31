from openpyxl import load_workbook
import csv

def puntoBase(archivoLectura, archivoEscritura):
        
    # Hoja de trabajo de operación
    #filesheet = "conciliacion.xlsx"
    wb = load_workbook(archivoEscritura)

    # Crear libro de trabajo, primero
    sheet = wb.worksheets[0]

    a=1
    b=4
    i=0

    with open(archivoLectura) as File:
        reader = csv.reader(File)
        for i in range (0, 8):
            next(reader)
        
        celda = sheet.cell(row=a, column=b)
        while celda.value:
            a=a+1
            celda = sheet.cell(row=a, column=b)
        
        for row in reader:
            print(row[4])
            dato = float(row[4])
            sheet.cell(row=a, column=b).value = str(dato)
            a=a+1


    wb.save(archivoEscritura)
